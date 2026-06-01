#!/usr/bin/env python3
"""Build a citation-audit Excel workbook from CSV rows.

The script intentionally does not search the web. It formats already-collected
evidence into a reviewer-friendly workbook and adds conservative next actions.
"""

from __future__ import annotations

import csv
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


STATUS_ORDER = ["◎ 真实", "○ 基本可信", "△ 待核", "重复", "疑似编造/不采用"]


ALIASES = {
    "编号": ["编号", "id", "number", "ref_id"],
    "原参考文献": ["原参考文献", "文献（著录）", "reference", "citation"],
    "最终判定": ["最终判定", "最终综合判定", "status", "判定"],
    "主证据URL": ["主证据URL", "证据/来源 URL", "source_url", "url"],
    "多源证据URL": ["多源证据URL", "secondary_urls", "extra_urls"],
    "截图文件": ["截图文件", "screenshot", "screenshot_file"],
    "匹配字段": ["匹配字段", "matched_fields"],
    "可靠性说明": ["可靠性说明", "综合依据", "evidence_note", "note"],
    "问题备注": ["问题备注", "issues", "problem_note"],
    "建议动作": ["建议动作", "next_action", "action"],
}


def usage() -> None:
    print("Usage: python build_audit_workbook.py input.csv output.xlsx", file=sys.stderr)
    raise SystemExit(2)


def normalize_status(value: str) -> str:
    text = (value or "").strip()
    if text.startswith("◎"):
        return "◎ 真实"
    if text.startswith("○") or "部分确认" in text or "基本可信" in text:
        return "○ 基本可信"
    if text.startswith("△") or "待核" in text or "建议核对" in text:
        return "△ 待核"
    if "重复" in text:
        return "重复"
    if "疑似" in text or "编造" in text or "不采用" in text:
        return "疑似编造/不采用"
    return text or "△ 待核"


def pick(row: dict[str, str], key: str) -> str:
    for name in ALIASES[key]:
        if name in row and row[name] is not None:
            return str(row[name]).strip()
    return ""


def next_action(status: str) -> str:
    if status == "◎ 真实":
        return "保留；如学校要求，可抽样补充数据库题录链接。"
    if status == "○ 基本可信":
        return "建议保留；定稿前补 CNKI/万方/维普/期刊官网/出版社页等强证据。"
    if status == "△ 待核":
        return "不要删除；继续人工核查数据库、期刊官网、机构页或引用反查。"
    if status == "重复":
        return "合并或删除重复编号，保留著录更完整的一条。"
    if status == "疑似编造/不采用":
        return "暂不采用；必须写明冲突证据或多源失败路径。"
    return "人工复核。"


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        raw = list(csv.DictReader(f))
    rows = []
    for item in raw:
        status = normalize_status(pick(item, "最终判定"))
        action = pick(item, "建议动作") or next_action(status)
        rows.append(
            {
                "编号": pick(item, "编号"),
                "原参考文献": pick(item, "原参考文献"),
                "最终判定": status,
                "主证据URL": pick(item, "主证据URL"),
                "多源证据URL": pick(item, "多源证据URL"),
                "截图文件": pick(item, "截图文件"),
                "匹配字段": pick(item, "匹配字段"),
                "可靠性说明": pick(item, "可靠性说明"),
                "问题备注": pick(item, "问题备注"),
                "建议动作": action,
            }
        )
    return rows


def style_sheet(ws, color: str = "1F4E78") -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def write_summary(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "给朋友看的结论"
    counts = Counter(row["最终判定"] for row in rows)
    ws.append(["项目", "结论/动作"])
    ws.append(["审查原则", "按身份核对，不按搜索排名核对；查不到不等于编造。"])
    ws.append(["最终分布", "；".join(f"{k} {counts.get(k, 0)} 条" for k in STATUS_ORDER)])
    ws.append(["优先动作", "先处理重复项，再人工核查△待核；○基本可信定稿前补强来源。"])
    ws.append(["删除原则", "只有明确元数据冲突或多源反查失败并有异常时，才考虑疑似编造/不采用。"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 110
    style_sheet(ws, "548235")


def write_details(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("最终核验明细")
    headers = list(rows[0].keys()) if rows else list(ALIASES.keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    widths = [10, 76, 16, 46, 46, 36, 34, 60, 50, 56]
    for idx, width in enumerate(widths[: len(headers)], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in [4, 5]:
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
    style_sheet(ws)


def write_actions(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("待处理清单")
    ws.append(["优先级", "编号", "最终判定", "下一步动作", "原参考文献"])
    priority = {"△ 待核": "P1", "重复": "P2", "○ 基本可信": "P3", "疑似编造/不采用": "P0"}
    selected = [r for r in rows if r["最终判定"] in priority]
    selected.sort(key=lambda r: priority[r["最终判定"]])
    for row in selected:
        ws.append([priority[row["最终判定"]], row["编号"], row["最终判定"], row["建议动作"], row["原参考文献"]])
    for idx, width in enumerate([10, 10, 16, 62, 86], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    style_sheet(ws, "9E480E")


def main(argv: list[str]) -> None:
    if len(argv) != 3:
        usage()
    input_path = Path(argv[1])
    output_path = Path(argv[2])
    rows = load_rows(input_path)
    if not rows:
        raise SystemExit("No rows found in input CSV")
    wb = Workbook()
    write_summary(wb, rows)
    write_details(wb, rows)
    write_actions(wb, rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main(sys.argv)
