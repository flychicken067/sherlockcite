#!/usr/bin/env python3
"""Build a citation-audit Excel workbook from CSV rows.

The script intentionally does not search the web. It formats already-collected
evidence into a reviewer-friendly workbook and adds conservative next actions.
Output sheets and statuses are English-first to avoid encoding issues in public
GitHub repositories and cross-platform terminal sessions.
"""

from __future__ import annotations

import csv
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


STATUS_ORDER = ["VERIFIED", "PLAUSIBLE", "PENDING", "DUPLICATE", "SUSPECT"]


ALIASES = {
    "id": ["id", "number", "ref_id"],
    "reference": ["reference", "citation", "raw_reference"],
    "status": ["status", "final_status", "decision"],
    "source_url": ["source_url", "url", "primary_url"],
    "secondary_urls": ["secondary_urls", "extra_urls", "multi_source_urls"],
    "screenshot_file": ["screenshot_file", "screenshot"],
    "matched_fields": ["matched_fields"],
    "evidence_note": ["evidence_note", "note", "reliability_note"],
    "problem_note": ["problem_note", "issues"],
    "next_action": ["next_action", "action"],
}


def usage() -> None:
    print("Usage: python build_audit_workbook.py input.csv output.xlsx", file=sys.stderr)
    raise SystemExit(2)


def normalize_status(value: str) -> str:
    text = (value or "").strip()
    upper = text.upper()
    if upper in STATUS_ORDER:
        return upper
    if "verified" in upper:
        return "VERIFIED"
    if "plausible" in upper:
        return "PLAUSIBLE"
    if "pending" in upper:
        return "PENDING"
    if "duplicate" in upper:
        return "DUPLICATE"
    if "suspect" in upper:
        return "SUSPECT"
    return text or "PENDING"


def pick(row: dict[str, str], key: str) -> str:
    for name in ALIASES[key]:
        if name in row and row[name] is not None:
            return str(row[name]).strip()
    return ""


def next_action(status: str) -> str:
    if status == "VERIFIED":
        return "Keep. Add a database or publisher URL only if the final submission requires it."
    if status == "PLAUSIBLE":
        return "Keep for now. Before final submission, add a stronger first-party/database record if possible."
    if status == "PENDING":
        return "Do not delete. Continue manual checks in databases, journal sites, institution pages, or exact citation trails."
    if status == "DUPLICATE":
        return "Merge or remove the duplicate entry; keep the most complete citation."
    if status == "SUSPECT":
        return "Do not use unless stronger evidence appears; document the conflict or failed source paths."
    return "Review manually."


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        raw = list(csv.DictReader(f))
    rows = []
    for item in raw:
        status = normalize_status(pick(item, "status"))
        action = pick(item, "next_action") or next_action(status)
        rows.append(
            {
                "id": pick(item, "id"),
                "reference": pick(item, "reference"),
                "status": status,
                "source_url": pick(item, "source_url"),
                "secondary_urls": pick(item, "secondary_urls"),
                "screenshot_file": pick(item, "screenshot_file"),
                "matched_fields": pick(item, "matched_fields"),
                "evidence_note": pick(item, "evidence_note"),
                "problem_note": pick(item, "problem_note"),
                "next_action": action,
            }
        )
    return rows


def style_sheet(ws, color: str = "1F4E78") -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border


def write_summary(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "Summary"
    counts = Counter(row["status"] for row in rows)
    ws.append(["Item", "Conclusion / Action"])
    ws.append(["Audit principle", "Verify by identity, not by search rank. Missing search results do not prove fabrication."])
    ws.append(["Status distribution", "; ".join(f"{k}: {counts.get(k, 0)}" for k in STATUS_ORDER)])
    ws.append(["Priority", "Handle duplicates, then manually check PENDING items. Add stronger evidence for PLAUSIBLE items before final submission."])
    ws.append(["Deletion rule", "Classify as SUSPECT only with explicit metadata conflict or repeated independent source failure plus abnormal citation structure."])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 120
    style_sheet(ws, "548235")


def write_details(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("Audit Details")
    headers = list(rows[0].keys()) if rows else list(ALIASES.keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    widths = [10, 80, 16, 48, 48, 36, 36, 64, 54, 64]
    for idx, width in enumerate(widths[: len(headers)], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in [4, 5]:
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"
    style_sheet(ws)


def write_actions(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("Action List")
    ws.append(["priority", "id", "status", "next_action", "reference"])
    priority = {"SUSPECT": "P0", "PENDING": "P1", "DUPLICATE": "P2", "PLAUSIBLE": "P3"}
    selected = [r for r in rows if r["status"] in priority]
    selected.sort(key=lambda r: priority[r["status"]])
    for row in selected:
        ws.append([priority[row["status"]], row["id"], row["status"], row["next_action"], row["reference"]])
    for idx, width in enumerate([10, 10, 16, 68, 90], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    style_sheet(ws, "9E480E")


def main(argv: list[str]) -> None:
    if len(argv) != 3:
        usage()
    input_path = Path(argv[1])
    output_path = Path(argv[2])
    rows = load_rows(input_path)
    if not rows:
        raise SystemExit("No rows found in input CSV")
    wb = Workbook()
    write_summary(wb, rows)
    write_details(wb, rows)
    write_actions(wb, rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main(sys.argv)
